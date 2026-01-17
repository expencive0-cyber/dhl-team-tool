"""Final AI Builder (Standard mapping).

Implements the core behavior from your notebook exports:
- Reads AF Input.xlsx (all sheets)
- Reads country code .xlsx (country list + DDP restrictions)
- Reads final AI template (header row 1 + constants from row 2)
- Writes an output workbook with one sheet per Source Sheet + _QC

Rules:
- Output schema locked to template headers and order
- Date uses Africa/Cairo, format DD-MM-YYYY
- Destination City & Destination Country are UPPERCASE
- Phone normalized to +E.164
- DDP = 'N' for countries listed in DDP sheet; otherwise 'Y'
- Missing email is NOT treated as an issue

This module is adapted from your 'AI DHL AF SINGLE' notebook. 
"""

from __future__ import annotations

from pathlib import Path
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

from .common import normalize_text, norm_key, trunc, today_str, only_digits

HIGHLIGHT_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

COUNTRY_ALIASES = {
    'IVORY COAST': 'COTE D IVOIRE',
    "COTE D'IVOIRE": 'COTE D IVOIRE',
    'COTE D’IVOIRE': 'COTE D IVOIRE',
    'REPUBLIC OF CONGO': 'CONGO',
    'DEMOCRATIC REPUBLIC OF CONGO': 'CONGO, THE DEMOCRATIC REPUBLIC OF',
    'DR CONGO': 'CONGO, THE DEMOCRATIC REPUBLIC OF',
    'RUSSIA': 'RUSSIAN FEDERATION, THE',
    'SOUTH KOREA': 'KOREA, REPUBLIC OF (SOUTH K.)',
    'NORTH KOREA': 'KOREA, THE D.P.R OF (NORTH K.)',
    'ESWATINI': 'SWAZILAND',
    'REPUBLIC OF SOUTH AFRICA': 'SOUTH AFRICA',
    'UAE': 'UNITED ARAB EMIRATES',
    'UK': 'UNITED KINGDOM',
    'USA': 'UNITED STATES OF AMERICA',
    'REUNION': 'REUNION, ISLAND OF',
    'SAO TOME AND PRINICIPE': 'SAO TOME AND PRINCIPE',
}

# Dial codes list is long in your notebook; we keep a practical subset.
# If a country isn't present, we will still output +<digits> as fallback.
DIAL_CODES = {
    'EGYPT': '20','UNITED ARAB EMIRATES':'971','UNITED KINGDOM':'44','UNITED STATES OF AMERICA':'1',
    'SAUDI ARABIA':'966','QATAR':'974','OMAN':'968','KUWAIT':'965','BAHRAIN':'973','JORDAN':'962',
    'MOROCCO':'212','TUNISIA':'216','ALGERIA':'213','NIGERIA':'234','KENYA':'254','SOUTH AFRICA':'27',
    'FRANCE':'33','GERMANY':'49','SPAIN':'34','ITALY':'39','TURKEY':'90',
    'INDIA':'91','PAKISTAN':'92','SINGAPORE':'65','JAPAN':'81','CHINA, PEOPLES REPUBLIC':'86'
}


def load_dhl_country_and_ddp(country_code_xlsx: Path):
    try:
        dhl_df = pd.read_excel(country_code_xlsx, sheet_name=0, engine='openpyxl')
    except Exception:
        dhl_df = pd.read_excel(country_code_xlsx, sheet_name='country code', engine='openpyxl')
    dhl_df = dhl_df.loc[:, ~dhl_df.columns.duplicated()].copy()

    code_col = name_col = None
    for c in dhl_df.columns:
        cu = str(c).upper()
        if 'COUNTRY CODE' in cu and code_col is None:
            code_col = c
        if 'COUNTRY NAME' in cu and name_col is None:
            name_col = c
    if code_col is None or name_col is None:
        raise RuntimeError('Could not detect DHL country code/name columns in country code .xlsx')

    dhl_df = dhl_df[[code_col, name_col]].copy()
    dhl_df.columns = ['DHL Country Code','DHL Country Name']
    dhl_df['key'] = dhl_df['DHL Country Name'].apply(norm_key)

    # DDP sheet: listed countries -> DDP=N
    try:
        ddp_df = pd.read_excel(country_code_xlsx, sheet_name='DDP', engine='openpyxl')
        ddp_df = ddp_df.loc[:, ~ddp_df.columns.duplicated()].copy()
    except Exception:
        ddp_df = pd.DataFrame()

    countries = []
    if not ddp_df.empty:
        for col in ddp_df.columns:
            for v in ddp_df[col].dropna().astype(str).tolist():
                vv = v.strip()
                if len(vv) >= 2 and any(ch.isalpha() for ch in vv):
                    if vv.upper() not in {'DDP','N','CONTENT','OB'}:
                        countries.append(vv)

    ddp_norm = set()
    for c in pd.Series(countries).drop_duplicates().tolist():
        ddp_norm.add(norm_key(COUNTRY_ALIASES.get(norm_key(c), c)))

    return dhl_df, ddp_norm


def map_country_to_dhl(country_raw: str, dhl_df: pd.DataFrame):
    c = normalize_text(country_raw)
    if not c:
        return None, None
    c_norm = norm_key(c)
    alias = COUNTRY_ALIASES.get(c_norm)
    c_key = norm_key(alias) if alias else c_norm

    row = dhl_df.loc[dhl_df['key'] == c_key]
    if not row.empty:
        r = row.iloc[0]
        return r['DHL Country Name'], r['DHL Country Code']

    # partial match fallback
    for _, r in dhl_df.iterrows():
        if c_key in r['key'] or r['key'] in c_key:
            return r['DHL Country Name'], r['DHL Country Code']

    return None, None


def ddp_flag(dhl_country_name: str | None, ddp_norm: set) -> str:
    if not dhl_country_name:
        return ''
    return 'N' if norm_key(dhl_country_name) in ddp_norm else 'Y'


def normalize_phone_e164(phone_raw: str, country_name: str) -> str:
    pr = normalize_text(phone_raw)
    if not pr:
        return ''

    # +...
    m = re.search(r'\+\s*([0-9][0-9\s\-\(\)]{6,})', pr)
    if m:
        return '+' + only_digits(m.group(1))

    digits = only_digits(pr)
    if not digits:
        return ''

    if digits.startswith('00'):
        return '+' + digits[2:]

    cname = normalize_text(country_name).upper()
    cname = COUNTRY_ALIASES.get(cname, cname)
    cc = DIAL_CODES.get(cname)

    # already includes country code
    if len(digits) >= 10 and not digits.startswith('0'):
        if cc and digits.startswith(cc):
            return '+' + digits
        # heuristic: treat as already international
        if len(digits) <= 15:
            return '+' + digits

    if cc:
        national = re.sub(r'^0+', '', digits)
        return '+' + cc + national

    # fallback
    return '+' + digits


def build_contacts_from_af(af_input_xlsx: Path):
    xls = pd.ExcelFile(af_input_xlsx)
    contacts_all = []

    header_map = {
        'DEPARTMENT': 'Department',
        'TITLE': 'Title',
        'GENDER': 'Gender',
        'FULL NAME': 'Full Name',
        'POSITION': 'Position',
        'COMPANY': 'Company',
        'CITY': 'City',
        'COUNTRY': 'Country',
        'LANGUAGE': 'Language',
        'STREET ADDRESS2': 'Street',
        'ADDRESS': 'Street',
        'STREET': 'Street',
        'TELEPHONE / MOBILE3': 'Phone',
        'TELEPHONE': 'Phone',
        'MOBILE': 'Phone',
        'POSTAL CODE': 'Postcode',
        'ZIP': 'Postcode',
        'EMAIL': 'Email',
    }

    skip_sheets = {'ALL DEPARTMENTS', 'LANGUAGE'}

    for sheet in xls.sheet_names:
        if norm_key(sheet) in {norm_key(s) for s in skip_sheets}:
            continue
        try:
            df = pd.read_excel(af_input_xlsx, sheet_name=sheet, engine='openpyxl')
        except Exception:
            continue
        if df.empty:
            continue
        df = df.loc[:, ~df.columns.duplicated()].copy()

        new_cols = []
        for c in df.columns:
            cu = norm_key(c)
            mapped = None
            for k, v in header_map.items():
                if k in cu:
                    mapped = v
                    break
            new_cols.append(mapped if mapped else c)
        df.columns = new_cols
        df = df.loc[:, ~df.columns.duplicated()].copy()

        keep_cols = ['Department','Title','Gender','Full Name','Position','Company',
                     'City','Country','Language','Street','Phone','Postcode','Email']
        for kc in keep_cols:
            if kc not in df.columns:
                df[kc] = np.nan
        df2 = df[keep_cols].copy()

        for c in ['Full Name','Company','Email','Country']:
            df2[c] = df2[c].apply(normalize_text)

        df2 = df2[(df2['Country'] != '') & ((df2['Full Name'] != '') | (df2['Company'] != ''))]
        if not df2.empty:
            df2['Source Sheet'] = sheet
            contacts_all.append(df2)

    if contacts_all:
        return pd.concat(contacts_all, ignore_index=True)

    return pd.DataFrame(columns=['Department','Title','Gender','Full Name','Position','Company',
                                 'City','Country','Language','Street','Phone','Postcode','Email','Source Sheet'])


def run_final_ai_standard(af_input_xlsx: Path, country_code_xlsx: Path, template_xlsx: Path, out_xlsx: Path):
    af_input_xlsx = Path(af_input_xlsx)
    country_code_xlsx = Path(country_code_xlsx)
    template_xlsx = Path(template_xlsx)
    out_xlsx = Path(out_xlsx)

    for p in (af_input_xlsx, country_code_xlsx, template_xlsx):
        if not p.exists():
            raise FileNotFoundError(f'Missing file: {p}')

    dhl_df, ddp_norm = load_dhl_country_and_ddp(country_code_xlsx)
    contacts = build_contacts_from_af(af_input_xlsx)

    wb_template = load_workbook(template_xlsx)
    ws_template = wb_template.active

    template_headers = [normalize_text(c.value) for c in next(ws_template.iter_rows(min_row=1, max_row=1))]
    second_row_values = [c.value for c in next(ws_template.iter_rows(min_row=2, max_row=2))]

    computed_cols = {
        'Order Number','Date','To Name',
        'Destination Building','Destination Street','Destination Suburb','Destination City',
        'Destination Postcode','Destination State','Destination Country',
        'Destination Email','Destination Phone','Company','Country Code','DDP'
    }

    constants = {}
    for h, v in zip(template_headers, second_row_values):
        if v is not None and str(v).strip() != '' and h not in computed_cols:
            constants[h] = v

    header_index = {h: i+1 for i, h in enumerate(template_headers)}
    date_str = today_str()

    wb_out = Workbook()
    if 'Sheet' in wb_out.sheetnames:
        del wb_out['Sheet']

    qc_rows = []
    total_highlighted = 0

    if contacts.empty:
        qc_ws = wb_out.create_sheet('_QC')
        qc_headers = ['Order Number','Source Sheet','Name','Email','Phone Raw','Phone E164','Country (raw)','DHL Country','DHL Code','DDP','Issues']
        qc_ws.append(qc_headers)
        wb_out.save(out_xlsx)
        return {'rows': 0, 'highlighted': 0, 'qc_rows': 0}

    for source_sheet, sheet_data in contacts.groupby('Source Sheet'):
        ws = wb_out.create_sheet(title=str(source_sheet)[:31])

        # headers
        for j, h in enumerate(template_headers, start=1):
            ws.cell(row=1, column=j, value=h)
        # constants row
        for h, col_idx in header_index.items():
            if h in constants:
                ws.cell(row=2, column=col_idx, value=constants[h])

        order_no = 1
        highlighted_sheet = 0

        for _, row in sheet_data.iterrows():
            issues = []

            title = normalize_text(row.get('Title',''))
            full_name = normalize_text(row.get('Full Name',''))
            tkey = title.replace('.', '').strip().upper()
            if tkey in {'MR','MRS','MS','DR'} and full_name:
                to_name = f"{title.replace('.', '').strip().title()} {full_name}".strip()
            else:
                to_name = full_name if full_name else normalize_text(row.get('Company',''))
            if not to_name:
                issues.append('Missing name and company')

            company = normalize_text(row.get('Company',''))
            country_raw = normalize_text(row.get('Country',''))
            dhl_name, dhl_code = map_country_to_dhl(country_raw, dhl_df)
            if not dhl_name:
                issues.append(f"Unknown country: '{country_raw}'")
            ddp = ddp_flag(dhl_name, ddp_norm) if dhl_name else ''

            street_raw = normalize_text(row.get('Street',''))
            street_parts = street_raw.split(',', 1)
            dest_building = trunc(street_parts[0].strip() if street_parts else '')
            dest_street = trunc(street_parts[1].strip() if len(street_parts) > 1 else street_raw)

            city = normalize_text(row.get('City',''))
            postcode = normalize_text(row.get('Postcode',''))
            if not postcode:
                m = re.search(r"\b\d{5}(-\d{4})?\b|\b[A-Z]{1,2}\d[A-Z\d]? ?\d[A-Z]{2}\b|\b\d{4,6}\b", street_raw, re.I)
                if m:
                    postcode = m.group(0).strip()

            if not street_raw:
                issues.append('Missing street')
            if not city:
                issues.append('Missing city')

            email = normalize_text(row.get('Email',''))
            phone_raw = normalize_text(row.get('Phone',''))
            phone_e164 = normalize_phone_e164(phone_raw, dhl_name or country_raw)
            if not phone_e164:
                issues.append('Missing phone')

            rec = {h: '' for h in template_headers}
            rec.update(constants)
            rec['Order Number'] = order_no
            rec['Date'] = date_str
            rec['To Name'] = to_name
            rec['Destination Building'] = dest_building
            rec['Destination Street'] = dest_street
            rec['Destination Suburb'] = ''
            rec['Destination City'] = city.upper()
            rec['Destination Postcode'] = postcode
            rec['Destination State'] = ''
            rec['Destination Country'] = (dhl_name if dhl_name else country_raw).upper()
            rec['Destination Email'] = email
            rec['Destination Phone'] = phone_e164
            rec['Company'] = company
            rec['Country Code'] = dhl_code if dhl_code else ''
            rec['DDP'] = ddp

            highlight = bool(issues)
            if highlight:
                highlighted_sheet += 1

            out_row = 2 + order_no
            for h, col_idx in header_index.items():
                ws.cell(row=out_row, column=col_idx, value=rec.get(h, ''))
            if highlight:
                for c in range(1, len(template_headers)+1):
                    ws.cell(row=out_row, column=c).fill = HIGHLIGHT_FILL

            qc_rows.append({
                'Order Number': order_no,
                'Source Sheet': str(source_sheet),
                'Name': to_name,
                'Email': email,
                'Phone Raw': phone_raw,
                'Phone E164': phone_e164,
                'Country (raw)': country_raw,
                'DHL Country': dhl_name or '',
                'DHL Code': dhl_code or '',
                'DDP': ddp,
                'Issues': '; '.join(issues) if issues else ''
            })

            order_no += 1

        total_highlighted += highlighted_sheet

    # QC sheet
    qc_ws = wb_out.create_sheet('_QC')
    qc_headers = ['Order Number','Source Sheet','Name','Email','Phone Raw','Phone E164','Country (raw)','DHL Country','DHL Code','DDP','Issues']
    qc_ws.append(qc_headers)
    for r in qc_rows:
        qc_ws.append([r.get(h,'') for h in qc_headers])

    wb_out.save(out_xlsx)
    return {'rows': len(qc_rows), 'highlighted': total_highlighted, 'qc_rows': len(qc_rows)}
